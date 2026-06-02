import os
import tempfile
import unittest

from openpyxl import Workbook

import origin_processor
import tensile_processor


class TensileSummaryExtractionTests(unittest.TestCase):
    def _save_workbook(self, workbook):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = os.path.join(temp_dir.name, "tensile.xlsx")
        workbook.save(path)
        return path

    def test_extracts_legacy_sheet1_layout(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append([
            "试样编号", "厚度\nmm", "宽度\nmm", "引伸计标距\nmm",
            "弹性模量\nGPa", "最大力\nN", "规定塑性延伸强度0.2%\nMPa",
            "抗拉强度\nMPa", "最大力塑性延伸率\n%", "最大力总延伸率\n%",
            "断后伸长率\n%", "断裂总延伸率\n%",
        ])
        ws.append([
            "N-1", "1.147", "12.47", "50.00", "186", "27160.02",
            "1534.5", "1899.3", "3.78", "4.60", "5.26", "6.05",
        ])

        _, groups = tensile_processor.extract_from_excel(self._save_workbook(wb))

        self.assertEqual(
            groups,
            {
                "N": [{
                    "id_num": "1",
                    "thick": "1.147",
                    "Rp": 1534,
                    "Rm": 1899,
                    "Ag": 3.8,
                    "At": 6.0,
                    "A": 5.3,
                    "has_note": False,
                }],
            },
        )

    def test_extracts_new_experiment_report_layout_by_headers(self):
        wb = Workbook()
        photos = wb.active
        photos.title = "实验前后照片"
        photos["C4"] = "实验前照片"

        ws = wb.create_sheet("实验报告")
        ws.append([
            "试样信息", "序号", "试样编号", "试样宽度bo", "试样厚度ao",
            "引伸计标距Le", "弹性模量E", "最大力Fm", "规定塑性延伸εp",
            "规定塑性延伸强度Rp", "抗拉强度Rm", "最大力塑性延伸率Ag",
            "最大力总延伸率Agt", "断后伸长率A", "断裂总延伸率At",
        ])
        ws.append([None, "单位", None, "mm", "mm", "mm", "MPa", "N", "%", "MPa", "MPa", "%", "%", "%", "%"])
        ws.append([
            "零件取样", None, "2-基材-H-1", 12.36, 0.995, 50, 200052,
            24744, 0.2, 1373, 2012, 4.04, 5.04, 4.96, 5.96,
        ])

        _, groups = tensile_processor.extract_from_excel(self._save_workbook(wb))

        self.assertEqual(
            groups,
            {
                "2-基材-H": [{
                    "id_num": "1",
                    "thick": "0.995",
                    "Rp": 1373,
                    "Rm": 2012,
                    "Ag": 4.0,
                    "At": 6.0,
                    "A": 5.0,
                    "has_note": False,
                }],
            },
        )


class TensileCurveSheetSelectionTests(unittest.TestCase):
    def _save_workbook(self, sheets):
        wb = Workbook()
        wb.remove(wb.active)
        for title, headers in sheets:
            ws = wb.create_sheet(title)
            ws.append(headers)
            ws.append([0] * len(headers))

        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = os.path.join(temp_dir.name, "curves.xlsx")
        wb.save(path)
        return path

    def test_selects_legacy_curve_sheet(self):
        path = self._save_workbook([
            ("实验前后照片", ["实验前", "实验后"]),
            ("曲线数据(1)", ["应力01", "应变01", "应力02", "应变02"]),
        ])

        self.assertEqual(origin_processor.select_tensile_curve_sheet(path), "曲线数据(1)")

    def test_selects_new_raw_data_sheet(self):
        path = self._save_workbook([
            ("实验前后照片", ["实验前", "实验后"]),
            ("原始数据", ["应力_1", "应变_1", "应力_2", "应变_2"]),
        ])

        self.assertEqual(origin_processor.select_tensile_curve_sheet(path), "原始数据")

    def test_rejects_workbooks_without_stress_strain_pairs(self):
        path = self._save_workbook([
            ("实验前后照片", ["实验前", "实验后"]),
            ("实验报告", ["试样编号", "抗拉强度Rm"]),
        ])

        with self.assertRaisesRegex(ValueError, "拉伸曲线数据"):
            origin_processor.select_tensile_curve_sheet(path)


if __name__ == "__main__":
    unittest.main()
